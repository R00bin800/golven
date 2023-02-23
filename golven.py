import matplotlib.pyplot as plt
import pandas as pd
import numpy as np
import statistics as stat
import os
#import xlsxwriter
import openpyxl as xl
import datetime
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors


#pi = np.pi

Hs_list = []

# location of file
home = r"C:\Users\robin\OneDrive - Windesheim Office365\school\Windesheim\Major 4\Waterbouwkunde 1\Golven auto"

output = home + r"\output_data"

print(output)


if not os.path.exists(output):
    os.makedirs(output)

df = pd.read_csv(home + r"\golven.csv", header=None, sep=", ", engine='python').transpose()
df.columns = ["Golfhoogte", "Golfperiode"]
g_hoogte_list = df["Golfhoogte"].tolist()
g_periode_list = df["Golfperiode"].tolist()


# df = pd.read_excel(path, header=None)           ##Importing the excel sheet from path, first row is data
# df.columns = ["temp"]                           
# df = df.temp.str.split(", ", expand=True)       ##Splitting all data from one cell to one per cell
# df=df.T                                         ##Transposing data to have correct Column/row
# df.columns=["Golfhoogte","Golfperiode"]         ##Giving Columns names
# #df.insert(0, "index", np.arange(df.shape[0]) )
# print(df)


# g_hoogte_list = list(map(int, df["Golfhoogte"].tolist()))       #making  hoogte collumn a list
# g_periode_list = list(map(int, df["Golfperiode"].tolist()))     #making  periode collumn a list
#print(g_hoogte_list)
#print(g_periode_list)

height_average = stat.fmean(g_hoogte_list)                      #get the average height from the hoogte list
period_average = stat.fmean(g_periode_list)                     #get the average period from the periode list

Hs_list = [i for i in g_hoogte_list if i > np.percentile(g_hoogte_list, 66.666667)]
Hs = stat.fmean(Hs_list)                                        #get the average height from the Hs_list, this is the significant wave height
#print(Hs_list)
#print(len(Hs_list))

index = [i for i, j in enumerate(g_hoogte_list) if j in Hs_list]          #get the location of the significant waves in the orignial wave list.
g_period_Hs = [g_periode_list[i] for i in index]    
Tp = stat.fmean(g_period_Hs)                                    #Find the average wave period for the significant waves


max_height = max(g_hoogte_list)                                 #get the highest wave and its corresponding peroid
max_corresponding_period = g_periode_list[g_hoogte_list.index(max_height)]

min_height = min(g_hoogte_list)                                 #get the lowest wave and its corresponding peroid
min_corresponding_period = g_periode_list[g_hoogte_list.index(min_height)]

longest_period = max(max_corresponding_period, min_corresponding_period)        #find the longest period between the lowest and highest wave


x = np.arange(0,longest_period,0.1)                                             #store the values of 0.1 steps between 0 and the longest_period
y = 0.5*max_height*np.sin((2*np.pi/max_corresponding_period)*x)                 #plot the highest wave
z = 0.5*min_height*np.sin((2*np.pi/min_corresponding_period)*x)                 #plot the lowest wave
w = y + z                                                                       #plot the sum

steps = np.arange(0, longest_period, step=np.pi / 2)
label = [str(j) + 'π' for j in np.arange(0, steps.size * 0.5, 0.5)]


print(f"significante golfhoogte: {Hs:.2f} m")
print(f"significante golfperiode: {Tp:.2f} sec")
print(f"hoogste golf: {max_height:.2f} m")
print(f"bijbehorende periode: {max_corresponding_period:.2f} s")
print(f"laagste golf: {min_height:.2f} m")
print(f"bijbehorende periode: {min_corresponding_period:.2f} s")            #print corresponding period in two decimals

#print("labels ", label)
plt.xlabel('time (s)')  # label on the x plane
plt.ylabel('height (m)') #label on the y plane
plt.title('Plot of wave height from 0 to %i seconds' %longest_period)   #title of the graph
plt.grid(True, which='both')                                            #show the grid
plt.plot(x,y,x,z,x,w)                                                   #plot everything for real
plt.gca().legend(('Highest wave', 'lowest wave', 'combined waves'))      #titles of each graph
plt.xticks(steps, label)                                                #show those value of pi
plt.savefig("myplot.png", dpi = 'figure')
#plt.show()
#print("steps ", steps)
#print(g_hoogte_list)
#print(Hs_list)
#print("answer: ")
#print(index)
#print(len(index))
#print(" ")


now=datetime.datetime.today().strftime("%Y-%m-%d_%H%M%S")



wb = xl.Workbook()
ws = wb.active

ws["A1"] = now
ws["O2"] = "Significante golfhoogte Hs ="
ws["P2"] = round(Hs,2)
ws["Q2"] = "m"
ws["O3"] = "Piekperiode Tp ="
ws["P3"] = round(Tp,2)
ws["Q3"] = "s"

ws["O7"] = "Hoogste golf ="
ws["P7"] = round(max_height,2)
ws["Q7"] = "m"
ws["O8"] = "Bijbehorende periode ="
ws["P8"] = round(max_corresponding_period,2)
ws["Q8"] = "s"

ws["O10"] = "Laagste golf ="
ws["P10"] = round(min_height,2)
ws["Q10"] = "m"
ws["O11"] = "Bijbehorende periode ="
ws["P11"] = round(min_corresponding_period,2)
ws["Q11"] = "s"

ws.cell("O25").font.italic = True
ws["O25"] = "Generated from code made by Robin van Marle"
# ws["O25"].font.italic = True

img = xl.drawing.image.Image('myplot.png')
ws.add_image(img, 'B2')

ws.column_dimensions['A'].width = 20
ws.column_dimensions['O'].width = 25
clr_background = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type="solid")

for row in ws["A1":"BA60"]: ##This will give you a cell range
    for cell in row:
        cell.fill = clr_background
wb.save(output + r"\\" +  "test" + ".xlsx")

